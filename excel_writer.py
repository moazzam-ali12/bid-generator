from __future__ import annotations

from typing import Dict, Any, List
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ── Brand colours ─────────────────────────────────────────────────────────────
NAVY      = "1E3347"
ORANGE    = "D4721A"
WHITE     = "FFFFFF"
ROW_LIGHT = "F4F6F9"
BORDER_C  = "DDE3EC"
SECTION_BG = "EEF1F5"


# ─── Shared Helpers ───────────────────────────────────────────────────────────

def _thin_border() -> Border:
    s = Side(style="thin", color=BORDER_C)
    return Border(left=s, right=s, top=s, bottom=s)


def _hfill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _fmt(val: Any) -> str:
    """Flatten any value to a clean Excel string."""
    if val is None:
        return ""
    if isinstance(val, list):
        parts = [_fmt(v) for v in val if v not in (None, "", [])]
        return "\n".join(f"• {p}" for p in parts) if parts else ""
    if isinstance(val, dict):
        return "\n".join(f"{k}: {_fmt(v)}" for k, v in val.items() if v)
    return str(val).strip()


def _autosize(ws: Worksheet, max_width: int = 60):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            v = cell.value
            if v:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), max_width)


# ─── Branded Title Block ──────────────────────────────────────────────────────

def _title_block(ws: Worksheet, title: str, meta: Dict, n_cols: int = 13) -> int:
    """Write navy title bar + orange accent + meta rows. Returns next available row."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value=title)
    c.font      = Font(bold=True, size=13, color=WHITE, name="Calibri")
    c.fill      = _hfill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.cell(row=2, column=1).fill = _hfill(ORANGE)
    ws.row_dimensions[2].height  = 4

    row = 3
    for label, value in [
        ("Project:",     meta.get("project", "")),
        ("Created By:",  meta.get("created_by", "")),
        ("Generated:",   meta.get("generated_date", "")),
    ]:
        if value:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=n_cols)
            lc = ws.cell(row=row, column=1, value=label)
            vc = ws.cell(row=row, column=5, value=str(value))
            lc.font = Font(bold=True, size=10, color=NAVY, name="Calibri")
            vc.font = Font(size=10, name="Calibri")
            lc.fill = vc.fill = _hfill(SECTION_BG)
            lc.alignment = vc.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

    return row + 1  # blank spacer row


# ─── Header + Data Row Writers ───────────────────────────────────────────────

def _header_row(ws: Worksheet, columns: List[str], row: int):
    for c, label in enumerate(columns, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font      = Font(bold=True, size=10, color=WHITE, name="Calibri")
        cell.fill      = _hfill(NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _thin_border()
    ws.row_dimensions[row].height = 22


def _data_rows(ws: Worksheet, columns: List[str], rows: List[Dict], start_row: int):
    for i, row_data in enumerate(rows):
        r    = start_row + i
        fill = _hfill(ROW_LIGHT) if i % 2 == 0 else _hfill(WHITE)
        for c, col in enumerate(columns, 1):
            val  = row_data.get(col, "")
            cell = ws.cell(row=r, column=c, value=_fmt(val))
            cell.font      = Font(size=10, name="Calibri")
            cell.fill      = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border    = _thin_border()
        ws.row_dimensions[r].height = 60


def _write_table_block(ws: Worksheet, tbl: Dict, meta: Dict,
                       col_widths: List[int] = None, n_cols: int = 13) -> int:
    """Write a complete titled table on an existing sheet. Returns next row after table."""
    title   = tbl.get("title", ws.title)
    columns = tbl.get("columns", [])
    rows    = tbl.get("rows", [])

    hdr_row = _title_block(ws, title, meta, n_cols=n_cols)
    _header_row(ws, columns, hdr_row)
    _data_rows(ws, columns, rows, hdr_row + 1)
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1).coordinate

    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    else:
        _autosize(ws)

    return hdr_row + len(rows) + 2


# ─── Cover Page Writer ────────────────────────────────────────────────────────

def _write_cover_page(ws: Worksheet, cover: Dict, meta: Dict):
    n_cols = 8
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value="Project Inspection and Testing Summary")
    c.font      = Font(bold=True, size=16, color=WHITE, name="Calibri")
    c.fill      = _hfill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.cell(row=2, column=1).fill = _hfill(ORANGE)
    ws.row_dimensions[2].height  = 6

    row = 3

    def _section_header(label: str):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
        c = ws.cell(row=row, column=1, value=label)
        c.font      = Font(bold=True, size=11, color=WHITE, name="Calibri")
        c.fill      = _hfill(ORANGE)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1

    def _info_row(label: str, value: str):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=n_cols)
        lc = ws.cell(row=row, column=1, value=label)
        vc = ws.cell(row=row, column=3, value=str(value or ""))
        lc.font = Font(bold=True, size=10, color=NAVY, name="Calibri")
        vc.font = Font(size=10, name="Calibri")
        lc.fill = vc.fill = _hfill(SECTION_BG)
        lc.alignment = vc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 18
        row += 1

    # ── Prepared By ──
    _section_header("PREPARED BY")
    _info_row("Name",          cover.get("created_by", ""))
    _info_row("Company",       cover.get("company", ""))
    _info_row("Phone",         cover.get("phone", ""))
    _info_row("Email",         cover.get("email", ""))
    row += 1

    # ── Project Information ──
    _section_header("PROJECT INFORMATION")
    _info_row("Project Name",    meta.get("project", ""))
    _info_row("Project Address", cover.get("project_address", ""))
    _info_row("County",          cover.get("county", ""))
    _info_row("City",            cover.get("city", ""))
    _info_row("Date Generated",  cover.get("date_run", meta.get("generated_date", "")))
    row += 1

    # ── Referenced Documents ──
    _section_header("REFERENCED DOCUMENTS")
    docs = cover.get("referenced_documents", [])
    if docs:
        for doc in docs:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
            c = ws.cell(row=row, column=1, value=f"• {doc}")
            c.font      = Font(size=10, name="Calibri")
            c.fill      = _hfill(ROW_LIGHT if row % 2 == 0 else WHITE)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = 18
            row += 1
    else:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
        ws.cell(row=row, column=1, value="No documents listed").font = Font(size=10, color="999999")
        row += 1
    row += 1

    # ── Tables Generated ──
    _section_header("TABLES GENERATED")
    tables = cover.get("tables_generated", [
        "Geotechnical Requirements",
        "Flatwork/Foundation Requirements",
        "Structural Requirements",
        "Quantity Estimation",
    ])
    for t in tables:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
        c = ws.cell(row=row, column=1, value=f"• {t}")
        c.font      = Font(size=10, name="Calibri")
        c.fill      = _hfill(ROW_LIGHT if row % 2 == 0 else WHITE)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 18
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    for i in range(3, n_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18


# ─── Quantity Estimation Writer ───────────────────────────────────────────────

def _write_quantities(ws: Worksheet, qty: Dict, meta: Dict):
    n_cols = 7
    hdr_row = _title_block(ws, "Quantity Estimation", meta, n_cols=n_cols)

    def _section_hdr(label: str) -> int:
        nonlocal hdr_row
        ws.merge_cells(start_row=hdr_row, start_column=1, end_row=hdr_row, end_column=n_cols)
        c = ws.cell(row=hdr_row, column=1, value=label)
        c.font      = Font(bold=True, size=11, color=WHITE, name="Calibri")
        c.fill      = _hfill(ORANGE)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[hdr_row].height = 20
        hdr_row += 1
        return hdr_row

    def _kv(label: str, value: Any):
        nonlocal hdr_row
        ws.merge_cells(start_row=hdr_row, start_column=1, end_row=hdr_row, end_column=3)
        ws.merge_cells(start_row=hdr_row, start_column=4, end_row=hdr_row, end_column=n_cols)
        lc = ws.cell(row=hdr_row, column=1, value=label)
        vc = ws.cell(row=hdr_row, column=4, value=_fmt(value))
        lc.font = Font(bold=True, size=10, color=NAVY, name="Calibri")
        vc.font = Font(size=10, name="Calibri")
        lc.fill = vc.fill = _hfill(SECTION_BG if hdr_row % 2 == 0 else WHITE)
        lc.alignment = vc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[hdr_row].height = 22
        hdr_row += 1

    # ── Section A — Lot Size ──
    _section_hdr("SECTION A — LOT SIZE")
    lot = qty.get("section_a_lot_size", {})
    _kv("Total Lot Size (sqft)",  lot.get("total_sqft"))
    _kv("Total Lot Size (acres)", lot.get("total_acres"))
    _kv("Source",                 lot.get("source"))
    hdr_row += 1

    # ── Section B — Foundations ──
    _section_hdr("SECTION B — FOUNDATION CONSTRUCTION ELEMENTS")
    foundations = qty.get("section_b_foundations", {})

    drilled = foundations.get("drilled_shafts", [])
    if drilled:
        _kv("Drilled Shafts", f"{len(drilled)} group(s)")
        for ds in drilled:
            _kv(f"  Count × Depth", f"{ds.get('count','?')} shafts @ {ds.get('depth','?')}")

    footings = foundations.get("spread_footings", [])
    if footings:
        _kv("Spread Footings", f"{len(footings)} size(s)")
        for sf in footings:
            label = sf.get("size_label", "")
            val   = (f"{sf.get('count','?')} ea | "
                     f"{sf.get('width','?')} × {sf.get('length','?')} × {sf.get('thickness','?')}")
            _kv(f"  {label}", val)

    lin = foundations.get("linear_footings", [])
    if lin:
        _kv("Linear Footings", f"{len(lin)} type(s)")
        for lf in lin:
            val = (f"{lf.get('count','?')} ea, L={lf.get('length','?')}, "
                   f"W={lf.get('width','?')}, D={lf.get('depth','?')}")
            _kv("  Linear Footing", val)

    if foundations.get("conflicts"):
        _kv("⚠ Conflicts", foundations["conflicts"])
    hdr_row += 1

    # ── Section C — Flatwork ──
    _section_hdr("SECTION C — TOTAL FLATWORK CONSTRUCTION ELEMENTS")
    flat = qty.get("section_c_flatwork", {})
    _kv("Total Pavement (sqft)",         flat.get("total_pavement_sqft"))
    _kv("Total Foundation Floor (sqft)", flat.get("total_foundation_floor_sqft"))
    _kv("Total Building (sqft)",         flat.get("total_building_sqft"))
    _kv("Source",                        flat.get("source"))
    if flat.get("conflicts"):
        _kv("⚠ Conflicts", flat["conflicts"])
    hdr_row += 1

    # ── Section D — Utilities ──
    _section_hdr("SECTION D — TOTAL UTILITIES WORK")
    util = qty.get("section_d_utilities", {})
    _kv("Water Line (LF)",     util.get("water_line_lf"))
    _kv("Storm Sewer (LF)",    util.get("storm_sewer_lf"))
    _kv("Sanitary Sewer (LF)", util.get("sanitary_sewer_lf"))
    _kv("Utility Depth",       util.get("utility_depth"))
    _kv("Source",              util.get("source"))
    hdr_row += 1

    # ── Section E — Structural Elements ──
    _section_hdr("SECTION E — TOTAL STRUCTURAL ELEMENTS")
    s_cols = ["Structural Element", "Material Type", "Quantity", "Unit",
              "Calculation Basis", "Source"]
    _header_row(ws, s_cols, hdr_row)
    hdr_row += 1

    struct_items = qty.get("section_e_structural", [])
    col_map = {
        "Structural Element": "structural_element",
        "Material Type":      "material_type",
        "Quantity":           "quantity",
        "Unit":               "unit",
        "Calculation Basis":  "calculation_basis",
        "Source":             "source",
    }
    for i, item in enumerate(struct_items):
        fill = _hfill(ROW_LIGHT) if i % 2 == 0 else _hfill(WHITE)
        for c, col in enumerate(s_cols, 1):
            val  = item.get(col_map[col], "")
            cell = ws.cell(row=hdr_row, column=c, value=_fmt(val))
            cell.font      = Font(size=10, name="Calibri")
            cell.fill      = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border    = _thin_border()
        ws.row_dimensions[hdr_row].height = 40
        hdr_row += 1

    # Conflicts
    conflicts = qty.get("conflicts", [])
    if conflicts:
        hdr_row += 1
        _section_hdr("CONFLICTS / VALIDATION ISSUES")
        for conflict in conflicts:
            ws.merge_cells(start_row=hdr_row, start_column=1, end_row=hdr_row, end_column=n_cols)
            c = ws.cell(row=hdr_row, column=1, value=f"⚠ {conflict}")
            c.font      = Font(size=10, color="CC0000", name="Calibri")
            c.fill      = _hfill("FFF3F3")
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[hdr_row].height = 30
            hdr_row += 1

    # Column widths
    widths = [28, 20, 14, 14, 30, 24, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─────────────────────────────────────────────────────────────────────────────
# build_workbook_v2 — 6-tab branded workbook (Herman's spec)
# ─────────────────────────────────────────────────────────────────────────────

def build_workbook_v2(all_tables: Dict[str, Any]) -> bytes:
    """
    Builds a branded workbook per Herman's spec:
    Cover Page → T1 Geotech → T2 Flatwork-Foundation → T3 Structural → Quantities → Gaps
    """
    wb = Workbook()
    wb.remove(wb.active)

    meta   = all_tables.get("meta", {})
    header = all_tables.get("header", {})
    cover  = all_tables.get("cover_page", {})

    # ── Tab 1: Cover Page ──
    ws_cover = wb.create_sheet("Cover Page")
    _write_cover_page(ws_cover, cover, meta)

    # ── Tab 2: T1 Geotech ──
    ws1 = wb.create_sheet("T1 – Geotech")
    tbl1 = all_tables.get("table1") or {
        "title": "Table 1 – Geotechnical Technical Requirements",
        "columns": ["Note"],
        "rows": [{"Note": "No data extracted."}],
    }
    # Add project site info block above table
    n_cols1 = 11
    row = _title_block(ws1, tbl1.get("title", "Table 1"), meta, n_cols=n_cols1)

    if header:
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols1)
        sh = ws1.cell(row=row, column=1, value="PROJECT SITE INFORMATION")
        sh.font      = Font(bold=True, size=10, color=WHITE, name="Calibri")
        sh.fill      = _hfill(ORANGE)
        sh.alignment = Alignment(horizontal="center", vertical="center")
        row += 1
        for label, value in [
            ("Project Address", header.get("project_address", "")),
            ("County",          header.get("county", "")),
            ("City",            header.get("city", "")),
            ("Referenced Docs", ", ".join(header.get("referenced_documents", []))),
        ]:
            ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            ws1.merge_cells(start_row=row, start_column=4, end_row=row, end_column=n_cols1)
            lc = ws1.cell(row=row, column=1, value=label)
            vc = ws1.cell(row=row, column=4, value=str(value or ""))
            lc.font = Font(bold=True, size=10, color=NAVY, name="Calibri")
            vc.font = Font(size=10, name="Calibri")
            lc.fill = vc.fill = _hfill(SECTION_BG)
            lc.alignment = vc.alignment = Alignment(wrap_text=True, vertical="top")
            ws1.row_dimensions[row].height = 18
            row += 1
        row += 1

    _header_row(ws1, tbl1.get("columns", []), row)
    _data_rows(ws1, tbl1.get("columns", []), tbl1.get("rows", []), row + 1)
    ws1.freeze_panes = ws1.cell(row=row + 1, column=1).coordinate
    for i, w in enumerate([22, 18, 20, 14, 16, 16, 18, 28, 28, 28, 20], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Tab 3: T2 Flatwork-Foundation ──
    ws2 = wb.create_sheet("T2 – Flatwork-Foundation")
    tbl2 = all_tables.get("table2") or {
        "title": "Table 2 – Flatwork/Foundation Technical Requirements",
        "columns": ["Note"],
        "rows": [{"Note": "No data extracted."}],
    }
    _write_table_block(ws2, tbl2, meta,
                       col_widths=[22, 16, 14, 12, 14, 16, 16, 16, 28, 20, 20, 24, 20],
                       n_cols=13)

    # ── Tab 4: T3 Structural ──
    ws3 = wb.create_sheet("T3 – Structural")
    tbl3 = all_tables.get("table3") or {
        "title": "Table 3 – Structural Technical Requirements",
        "columns": ["Note"],
        "rows": [{"Note": "No data extracted."}],
    }
    _write_table_block(ws3, tbl3, meta,
                       col_widths=[24, 18, 20, 24, 20, 28, 22],
                       n_cols=7)

    # ── Tab 5: Quantity Estimation ──
    ws_qty = wb.create_sheet("Quantity Estimation")
    _write_quantities(ws_qty, all_tables.get("quantity_estimation", {}), meta)

    # ── Tab 6: Gaps & Assumptions ──
    gaps = all_tables.get("assumptions_or_gaps") or []
    if gaps:
        ws_g = wb.create_sheet("Gaps & Assumptions")
        ws_g.merge_cells("A1:D1")
        c = ws_g["A1"]
        c.value     = "Assumptions & Gaps Identified by AI"
        c.font      = Font(bold=True, size=12, color=WHITE, name="Calibri")
        c.fill      = _hfill(NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws_g.row_dimensions[1].height = 26
        for i, note in enumerate(gaps, 2):
            cell = ws_g.cell(row=i, column=1, value=f"• {note}")
            cell.font      = Font(size=10, name="Calibri")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill      = _hfill(ROW_LIGHT if i % 2 == 0 else WHITE)
            ws_g.row_dimensions[i].height = 40
        ws_g.column_dimensions["A"].width = 120

    wb.active = wb.worksheets[0]
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()